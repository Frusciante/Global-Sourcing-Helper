import os
import time
import json
import shutil
import subprocess
import requests
import xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from tkinter import messagebox

# [중요] 외부 라이브러리
import google.generativeai as genai 
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException, TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

class SourcingProcessor:
    def __init__(self, config, log_callback):
        self.config = config
        self.log_callback = log_callback
        self.is_running = True
        self.target_file = self.config.get('EXCEL_FILE', 'windly-excel-bulk-upload-ver9.xlsx')
        
        # API 키 설정
        raw_keys = self.config['GEMINI_API_KEY']
        self.api_keys = [k.strip() for k in raw_keys.split(',') if k.strip()]
        self.current_key_idx = 0
        
        # 모델 설정
        self.model_candidates = [
            "gemini-2.5-flash", 
            "gemini-2.5-flash-lite", 
            "gemini-2.0-flash",
            "gemini-1.5-flash"
        ]
        self.current_model_idx = 0
        self.model = None 
        self.proc = None

        self.log_callback("📋 [초기화] 설정 로드 및 카테고리 데이터 준비 중...")
        self.load_categories_from_excel()

        try:
            self._configure_genai()
        except Exception as e:
            self.log_callback(f"❌ [Error] Gemini 초기 설정 실패: {e}")

    # ==========================
    # 1. AI 설정 및 로테이션
    # ==========================
    def _configure_genai(self):
        if not self.api_keys: 
            self.log_callback("❌ [Config] API 키가 없습니다.")
            return
        current_key = self.api_keys[self.current_key_idx]
        masked_key = f"{current_key[:5]}...{current_key[-5:]}"
        
        try:
            genai.configure(api_key=current_key)
            model_name = self.model_candidates[self.current_model_idx]
            self.log_callback(f"🔑 [AI] 키 적용 완료 ({self.current_key_idx + 1}/{len(self.api_keys)}) | 모델: {model_name}")
            self.model = genai.GenerativeModel(model_name)
        except Exception as e:
            self.log_callback(f"❌ [AI] 설정 오류: {e}")

    def _rotate_api_key(self):
        if len(self.api_keys) <= 1: 
            self.log_callback("⚠️ [AI] 교체할 여분 키가 없습니다.")
            return False
        self.current_key_idx = (self.current_key_idx + 1) % len(self.api_keys)
        self.log_callback("🔄 [AI] 한도 초과 감지! 다음 키로 교체합니다...")
        self._configure_genai()
        return True

    def _call_gemini_with_retry(self, prompt, context=""):
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if not self.model: self._configure_genai()
                response = self.model.generate_content(prompt)
                if response and response.text: 
                    return response.text.strip()
            except Exception as e:
                error_msg = str(e).lower()
                if "429" in error_msg or "quota" in error_msg:
                    self.log_callback(f"⏳ [AI] 429 Too Many Requests ({context}). 키 교체 시도.")
                    if self._rotate_api_key(): 
                        time.sleep(1)
                        continue
                    time.sleep(10)
                    continue
                elif "404" in error_msg:
                    self.log_callback(f"⚠️ [AI] 모델 오류. 모델 변경.")
                    self.current_model_idx = (self.current_model_idx + 1) % len(self.model_candidates)
                    self.model = genai.GenerativeModel(self.model_candidates[self.current_model_idx])
                    continue
                else:
                    self.log_callback(f"⚠️ [AI] 오류 발생: {error_msg}")
                    return None
        return None

    # ==========================
    # 2. 엑셀 및 데이터 처리
    # ==========================
    def load_categories_from_excel(self):
        try:
            if not os.path.exists(self.target_file): 
                self.log_callback(f"⚠️ [Excel] 파일 없음: {self.target_file}")
                return
            self.coupang_cat = pd.read_excel(self.target_file, sheet_name='쿠팡 전체 카테고리 (240517)')
            self.naver_cat = pd.read_excel(self.target_file, sheet_name='네이버 전체 카테고리 (251215)')
            self.log_callback(f"✅ [Excel] 카테고리 데이터 로드 완료")
        except Exception as e:
            self.log_callback(f"❌ [Excel] 로드 실패: {e}")

    def find_best_category(self, hint, platform='coupang'):
        df = self.coupang_cat if platform == 'coupang' else self.naver_cat
        if df is None: return ""
        target_col = '여기서 카테고리를 복사해주세요'
        keywords = hint.replace('>', ' ').split()
        for kw in reversed(keywords):
            if len(kw.strip()) < 2: continue
            match = df[df[target_col].str.contains(kw, na=False, case=False)]
            if not match.empty: return match.iloc[0][target_col]
        return ""

    # [신규 기능] 키워드 자체로 카테고리 결정 (일관성 유지)
    def determine_master_category(self, keyword):
        self.log_callback(f"🧠 [Category] '{keyword}'의 대표 카테고리 분석 중...")
        prompt = (
            f"검색어: '{keyword}'\n"
            f"이 검색어가 속할 가장 적절한 한국 이커머스 카테고리 경로를 하나만 추론해줘.\n"
            f"예시: 가구/인테리어 > 인테리어 조명 > 단스탠드\n"
            f"설명 없이 경로만 출력해."
        )
        cat_hint = self._call_gemini_with_retry(prompt, "카테고리 결정")
        
        if cat_hint:
            cp = self.find_best_category(cat_hint, 'coupang')
            nv = self.find_best_category(cat_hint, 'naver')
            self.log_callback(f"   ㄴ 결정됨: [쿠팡] {cp} / [네이버] {nv}")
            return cp, nv
        return "", ""

    def append_to_excel(self, data_row):
        try:
            wb = openpyxl.load_workbook(self.target_file)
            ws = wb['엑셀 수집 양식 (Ver.9)']
            
            start_row = 7
            while ws.cell(row=start_row, column=4).value is not None:
                start_row += 1
            
            tags_value = data_row['tags']
            if isinstance(tags_value, list):
                tags_value = ", ".join(tags_value)
            
            # [중요] data_row에 이미 고정된 카테고리가 들어있음
            ws.cell(row=start_row, column=2, value=data_row['cp_cat'])
            ws.cell(row=start_row, column=3, value=data_row['nv_cat'])
            ws.cell(row=start_row, column=4, value=data_row['title'])
            ws.cell(row=start_row, column=5, value=tags_value)
            ws.cell(row=start_row, column=6, value=data_row['url'])
            ws.cell(row=start_row, column=7, value=0)
            ws.cell(row=start_row, column=8, value='무료')
            ws.cell(row=start_row, column=9, value=0)
            ws.cell(row=start_row, column=10, value=5000)
            ws.cell(row=start_row, column=11, value=10000)
            ws.cell(row=start_row, column=12, value=data_row['manufacturer'])
            ws.cell(row=start_row, column=13, value=data_row['brand'])
            ws.cell(row=start_row, column=14, value=data_row['model'])
            
            wb.save(self.target_file)
            self.log_callback(f"💾 [Excel] {start_row}행 저장 | {data_row['title'][:10]}...")
        except Exception as e:
            self.log_callback(f"❌ [Excel] 저장 실패: {e}")

    # ==========================
    # 3. 브라우저 및 탐색
    # ==========================
    def init_driver(self):
        try:
            subprocess.run("taskkill /F /IM chrome.exe /T", shell=True, stderr=subprocess.DEVNULL)
            time.sleep(1)
        except: pass

        current_folder = os.getcwd()
        bot_profile_path = os.path.join(current_folder, "bot_profile")
        real_user_data = os.path.join(os.environ['LOCALAPPDATA'], 'Google', 'Chrome', 'User Data')

        if not os.path.exists(bot_profile_path):
            self.log_callback("♻️ [Init] 프로필 복제 중... (최초 1회)")
            try:
                shutil.copytree(real_user_data, bot_profile_path, ignore=shutil.ignore_patterns('*.lock', 'Singleton*', '*.tmp', 'Cache*', 'Code Cache*'))
            except: pass

        chrome_exe_path = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        if not os.path.exists(chrome_exe_path): chrome_exe_path = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        
        debug_port = 9222
        cmd = [
            chrome_exe_path,
            f"--remote-debugging-port={debug_port}",
            f"--user-data-dir={bot_profile_path}",
            "--profile-directory=Default",
            "--no-first-run", "--remote-allow-origins=*"
        ]
        
        self.log_callback(f"🚀 [Init] 크롬 프로세스 시작")
        self.proc = subprocess.Popen(cmd)
        time.sleep(3)

        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", f"127.0.0.1:{debug_port}")
        
        try:
            service = Service(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_options)
            self.log_callback("✅ [Init] Selenium 연결 성공")
            return driver
        except Exception as e:
            self.log_callback(f"❌ [Init] 연결 실패: {e}")
            raise e

    def detect_and_translate(self, url, html_source, keyword):
        try:
            target_lang = None
            if any(site in url for site in ['taobao', 'tmall', '1688']): target_lang = "중국어 간체"
            elif any(site in url for site in ['rakuten', 'yahoo']): target_lang = "일본어"
            elif any(site in url for site in ['amazon', 'ebay']): target_lang = "영어"

            if target_lang:
                self.log_callback(f"🌐 [Trans] 타겟 언어: {target_lang}")
                trans_prompt = f"쇼핑 검색어 '{keyword}'를 '{target_lang}'로 번역해줘. 단어만 출력."
                translated = self._call_gemini_with_retry(trans_prompt, "번역")
                if translated: 
                    self.log_callback(f"   ㄴ 번역: '{keyword}' -> '{translated}'")
                    return translated
            return keyword
        except Exception as e:
            self.log_callback(f"⚠️ [Trans] 번역 실패: {e}")
            return keyword

    def get_shopping_products(self, driver, url, keyword, count):
        while self.is_running:
            try:
                self.log_callback(f"🔍 [Search] '{keyword}' 검색 시작...")
                driver.get(url)
                time.sleep(3)

                # 1. 검색창 찾기
                search_input = None
                search_selectors = [
                    "input#twotabsearchtextbox", "input#q", "input[name='q']", 
                    "input[type='search']", "input[name='keyword']", "input#mq", "input[id*='search']"
                ]

                for sel in search_selectors:
                    try:
                        search_input = WebDriverWait(driver, 3).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                        )
                        if search_input: break
                    except: continue

                # 2. 검색 수행
                if search_input:
                    try:
                        search_input.clear()
                        search_input.send_keys(keyword)
                        time.sleep(1)
                        search_input.send_keys(Keys.ENTER)
                        self.log_callback("   ㄴ 검색어 입력 및 엔터 완료")
                        
                        time.sleep(2)
                        if driver.current_url == url: # 페이지 안 바뀌면 클릭 시도
                            btn_selectors = ["input[type='submit']", "button[class*='search']", "span[class*='search-icon']", "#nav-search-submit-button"]
                            for btn_sel in btn_selectors:
                                try:
                                    driver.find_element(By.CSS_SELECTOR, btn_sel).click()
                                    break
                                except: pass
                        time.sleep(5)
                    except Exception as e:
                        self.log_callback(f"❌ [Search] 입력 오류: {e}")
                        raise e
                else:
                    raise Exception("검색창 미발견")

                # 3. 상품 목록 수집
                selectors = ["[class*='title--']", "[class*='Title--']", "div.title", "div.item-name", "a[id*='item-title']", "h1", "h2", "h3", "span.a-text-normal"]
                products = []
                
                for selector in selectors:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if not elements: continue
                    
                    valid_elements = [el for el in elements if len(el.text.strip()) > 5]
                    
                    if valid_elements:
                        self.log_callback(f"   ㄴ 목록 발견: '{selector}' ({len(valid_elements)}개)")
                        for el in valid_elements:
                            product_name = el.text.strip()
                            product_link = driver.current_url 
                            
                            try:
                                if el.tag_name == 'a': product_link = el.get_attribute('href')
                                else:
                                    try: parent_a = el.find_element(By.XPATH, "./ancestor::a"); product_link = parent_a.get_attribute('href')
                                    except:
                                        try: child_a = el.find_element(By.TAG_NAME, "a"); product_link = child_a.get_attribute('href')
                                        except: pass
                            except: pass

                            if product_link == driver.current_url: continue 
                            products.append((product_name, product_link))
                        
                        if len(products) >= 3: break
                
                if not products: raise Exception("유효한 상품 없음")
                return products[:count]

            except WebDriverException as we:
                self.log_callback(f"🚨 [Browser] 연결 끊김 재시작: {we}")
                raise we
            except Exception as e:
                self.log_callback(f"⚠️ [Search] 검색 실패: {e}")
                return []

    def extract_full_info(self, p_name):
        prompt = (
            f"Analyze: '{p_name}'\n"
            "Is this a product? If navigational text, 'is_valid': false.\n"
            "Rules:\n"
            "1. brand: if unknown output 'NULL'.\n"
            "2. productTitle: Korean translation.\n"
            "JSON Output: { 'is_valid': true, 'productTitle': '...', 'manufacturer': '...', 'brand': '...', 'model': '...', 'keywords': [], 'category_hint': '...' }"
        )
        res = self._call_gemini_with_retry(prompt, "정보추출")
        if res:
            try:
                data = json.loads(res.replace('```json','').replace('```','').strip())
                if not data.get('is_valid', True): 
                    self.log_callback(f"   🗑️ [Filter] 유효하지 않은 상품 제외: {p_name[:10]}...")
                    return None
                return data
            except: return None
        return None

    def check_trademark(self, brand):
        if not brand or str(brand).upper() in ["NULL", "NONE", "N/A"]: return True
        api_url = "https://plus.kipris.or.kr/kipo-api/kipi/trademarkInfoSearchService/getWordSearch"
        params = {'searchString': brand, 'ServiceKey': self.config['KIPRIS_API_KEY']}
        try:
            res = requests.get(api_url, params=params, timeout=5)
            if res.status_code != 200: return True
            root = ET.fromstring(res.content)
            count = int(root.find(".//totalCount").text)
            if count > 0:
                self.log_callback(f"   ❌ [KIPRIS] 상표권 발견! '{brand}' ({count}건)")
                return False
            return True
        except: return True 

    # [핵심] 런 루프 변경
    def run(self):
        keywords = [k.strip() for k in self.config['TARGET_ITEMS'].split(",") if k.strip()]
        urls = [u.strip() for u in self.config['SHOP_URLS'].split(",") if u.strip()]
        max_count = int(self.config.get('ITEM_COUNT', 10))
        
        driver = self.init_driver()
        try:
            for kw in keywords:
                if not self.is_running: break
                self.log_callback(f"\n=== 🏁 키워드 작업 시작: {kw} ===")
                
                # 1. 여기서 카테고리를 고정합니다 (Master Category)
                fixed_cp_cat, fixed_nv_cat = self.determine_master_category(kw)
                
                for shop_url in urls:
                    if not self.is_running: break
                    try:
                        driver.get(shop_url)
                        time.sleep(3)
                        t_kw = self.detect_and_translate(shop_url, driver.page_source, kw)
                        product_list = self.get_shopping_products(driver, shop_url, t_kw, max_count)
                        
                        self.log_callback(f"📊 [Info] 총 {len(product_list)}개 상품 분석 시작...")
                        
                        for i, (p_name, p_url) in enumerate(product_list):
                            if not self.is_running: break
                            self.log_callback(f"🔎 [{i+1}/{len(product_list)}] 상세 분석 중...")
                            
                            info = self.extract_full_info(p_name)
                            time.sleep(5)
                            
                            if info:
                                if self.check_trademark(info['brand']):
                                    
                                    # [중요] 개별 상품 카테고리(hint)를 무시하고, 고정된 카테고리를 사용
                                    self.append_to_excel({
                                        'cp_cat': fixed_cp_cat, # 고정값 사용
                                        'nv_cat': fixed_nv_cat, # 고정값 사용
                                        'title': info['productTitle'], 
                                        'tags': info['keywords'],
                                        'url': p_url,
                                        'manufacturer': info['manufacturer'],
                                        'brand': info['brand'], 'model': info['model']
                                    })
                            time.sleep(1)
                    except Exception as e:
                        self.log_callback(f"⚠️ [Loop Error] {e}")
                        try: driver.quit(); self.proc.kill()
                        except: pass
                        driver = self.init_driver()
        finally:
            try: driver.quit(); self.proc.kill()
            except: pass
            self.log_callback("\n🏁 [Finish] 작업 종료")

    def stop(self):
        self.is_running = False
        self.log_callback("🛑 [Stop] 중지 요청됨")