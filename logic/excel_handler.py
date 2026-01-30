import os
import pandas as pd
import openpyxl

class ExcelHandler:
    def __init__(self, target_file, log_callback):
        self.target_file = target_file
        self.log_callback = log_callback
        self.coupang_cat = None
        self.naver_cat = None
        
        # 초기 로드
        self.load_categories()

    def load_categories(self):
        """카테고리 엑셀 파일 로드"""
        try:
            if not os.path.exists(self.target_file): 
                self.log_callback(f"⚠️ [Excel] 파일 없음: {self.target_file}")
                return
            # 데이터 타입을 str로 강제 변환하여 로드 (에러 방지)
            self.coupang_cat = pd.read_excel(self.target_file, sheet_name='쿠팡 전체 카테고리 (240517)', dtype=str)
            self.naver_cat = pd.read_excel(self.target_file, sheet_name='네이버 전체 카테고리 (251215)', dtype=str)
            self.log_callback(f"✅ [Excel] 카테고리 데이터 로드 완료")
        except Exception as e:
            self.log_callback(f"❌ [Excel] 로드 실패: {e}")

    def find_best_category(self, ai_path_hint, platform='coupang'):
        """
        [개선된 알고리즘]
        1. 단순 포함(in) 대신 단어 단위 분리(Split) 후 일치 여부 확인
        2. 경로의 '마지막 단어'가 정확히 일치하면 가산점 부여
        3. 점수가 같으면 '더 짧은 경로'를 선택 (군더더기 없는 매칭 선호)
        """
        df = self.coupang_cat if platform == 'coupang' else self.naver_cat
        if df is None or not ai_path_hint: return ""
        
        target_col = '여기서 카테고리를 복사해주세요'
        
        # 1. AI 힌트 전처리 (특수문자 제거 및 리스트화)
        # 예: "문구 > 필기구 > 연필" -> ['문구', '필기구', '연필']
        hint_keywords = [k.strip() for k in ai_path_hint.replace('>', ' ').split() if len(k.strip()) > 0]
        if not hint_keywords: return ""
        
        hint_last_word = hint_keywords[-1] # 핵심 키워드 (예: 연필)

        best_match = ""
        max_score = -1 # 초기값
        
        # 후보군 필터링 (속도 최적화: 핵심 단어가 포함된 것만 1차 조회)
        candidates = df[df[target_col].str.contains(hint_last_word, na=False, case=False)]
        if candidates.empty:
            candidates = df # 없으면 전체 검색

        for cat_path in candidates[target_col]:
            if not isinstance(cat_path, str): continue
            
            # 2. 카테고리 경로를 단어 단위로 쪼개기 (Tokenization)
            # 구분자(>, /)를 모두 공백으로 바꾸고 리스트로 만듦
            # 예: "문구/사무용품>연필꽂이" -> ['문구', '사무용품', '연필꽂이']
            cat_tokens = cat_path.replace('>', ' ').replace('/', ' ').split()
            cat_tokens = [t.strip() for t in cat_tokens if t.strip()]
            
            score = 0
            
            # [채점 기준 1] 단어 일치 개수 (Set Intersection 개념)
            # '연필'을 찾는데 '연필꽂이' 토큰은 '연필'과 다르므로 매칭되지 않음
            for kw in hint_keywords:
                if kw in cat_tokens: 
                    score += 10 # 단순 포함보다 높은 점수 부여
                elif kw in cat_path: 
                    score += 1  # (보조) 단어는 안 맞지만 글자가 포함되면 소폭 점수 (예: 띄어쓰기 차이)

            # [채점 기준 2] 마지막 단어(Leaf Category) 완전 일치 보너스 (핵심!)
            if cat_tokens and (cat_tokens[-1] == hint_last_word):
                score += 50 # 강력한 가산점 (확실한 타겟)

            # [갱신 로직]
            if score > max_score:
                max_score = score
                best_match = cat_path
            
            elif score == max_score:
                # [동점 처리 수정] 더 짧은 것을 선택!
                # 이유: '연필'(짧음)이 '연필 교정 그립'(김)보다 사용자의 의도에 가까울 확률이 높음 (일반화)
                if len(cat_path) < len(best_match):
                    best_match = cat_path

        return best_match

    def save_product(self, data_row):
        """수집된 상품 정보를 엑셀에 추가"""
        try:
            wb = openpyxl.load_workbook(self.target_file)
            ws = wb['엑셀 수집 양식 (Ver.9)']
            
            # 빈 행 찾기 (4열 상품명 기준)
            start_row = 7
            while ws.cell(row=start_row, column=4).value is not None:
                start_row += 1
            
            # 리스트 -> 문자열 변환
            tags_value = data_row['tags']
            if isinstance(tags_value, list):
                tags_value = ", ".join(tags_value)
            
            # 엑셀 쓰기
            ws.cell(row=start_row, column=2, value=data_row['cp_cat'])
            ws.cell(row=start_row, column=3, value=data_row['nv_cat'])
            ws.cell(row=start_row, column=4, value=data_row['title'])
            ws.cell(row=start_row, column=5, value=tags_value)
            ws.cell(row=start_row, column=6, value=data_row['url'])
            
            # 고정값들
            ws.cell(row=start_row, column=7, value=0)
            ws.cell(row=start_row, column=8, value='무료')
            ws.cell(row=start_row, column=9, value=0)
            ws.cell(row=start_row, column=10, value=5000)
            ws.cell(row=start_row, column=11, value=10000)
            ws.cell(row=start_row, column=12, value=data_row['manufacturer'])
            ws.cell(row=start_row, column=13, value=data_row['brand'])
            ws.cell(row=start_row, column=14, value=data_row['model'])
            
            wb.save(self.target_file)
            self.log_callback(f"💾 [Excel] {start_row}행 저장 | {data_row['title'][:10]}...")
            
        except Exception as e:
            self.log_callback(f"❌ [Excel] 저장 실패: {e}")